import openpyxl
def getRowcount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)
def getColumncount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)
def readData(file,sheetName,rowNumber,columnNumber):
    workbook=openpyxl.load_workbook(file,sheetName,rowNumber,columnNumber)
    sheet=workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNumber,column=columnNumber).value

def writeData(file,sheetName,rowNumber,columnNumber,data):
    workbook = openpyxl.load_workbook(file, sheetName, rowNumber, columnNumber)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNumber,column=columnNumber).value=data
    workbook.save(file)



